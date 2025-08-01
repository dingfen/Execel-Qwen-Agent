import os
import cv2
import numpy as np
from io import BytesIO
from typing import Any, Union
from PIL import Image, UnidentifiedImageError
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import pandas

from rapid_table_det_paddle.inference import TableDetector
from wired_table_rec.utils.utils import VisTable
from table_cls import TableCls
from wired_table_rec.main import WiredTableInput, WiredTableRecognition
from lineless_table_rec.main import LinelessTableInput, LinelessTableRecognition
from rapidocr import RapidOCR

root_dir = Path(__file__).resolve().parent
InputType = Union[str, np.ndarray, bytes, Path, Image.Image]


# from https://github.com/RapidAI/TableStructureRec/blob/main/wired_table_rec/utils/utils.py#L575
class LoadImage:
    def __init__(self):
        pass

    def __call__(self, img: InputType) -> np.ndarray:
        if not isinstance(img, InputType.__args__):
            raise LoadImageError(
                f"The img type {type(img)} does not in {InputType.__args__}"
            )

        origin_img_type = type(img)
        img = self.load_img(img)
        img = self.convert_img(img, origin_img_type)
        return img

    def load_img(self, img: InputType) -> np.ndarray:
        if isinstance(img, (str, Path)):
            self.verify_exist(img)
            try:
                img = self.img_to_ndarray(Image.open(img))
            except UnidentifiedImageError as e:
                raise LoadImageError(f"cannot identify image file {img}") from e
            return img

        if isinstance(img, bytes):
            img = self.img_to_ndarray(Image.open(BytesIO(img)))
            return img

        if isinstance(img, np.ndarray):
            return img

        if isinstance(img, Image.Image):
            return self.img_to_ndarray(img)

        raise LoadImageError(f"{type(img)} is not supported!")

    def img_to_ndarray(self, img: Image.Image) -> np.ndarray:
        if img.mode == "1":
            img = img.convert("L")
            return np.array(img)
        return np.array(img)

    def convert_img(self, img: np.ndarray, origin_img_type: Any) -> np.ndarray:
        if img.ndim == 2:
            return cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)

        if img.ndim == 3:
            channel = img.shape[2]
            if channel == 1:
                return cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)

            if channel == 2:
                return self.cvt_two_to_three(img)

            if channel == 3:
                if issubclass(origin_img_type, (str, Path, bytes, Image.Image)):
                    return cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
                return img

            if channel == 4:
                return self.cvt_four_to_three(img)

            raise LoadImageError(
                f"The channel({channel}) of the img is not in [1, 2, 3, 4]"
            )

        raise LoadImageError(f"The ndim({img.ndim}) of the img is not in [2, 3]")

    @staticmethod
    def cvt_two_to_three(img: np.ndarray) -> np.ndarray:
        """gray + alpha → BGR"""
        img_gray = img[..., 0]
        img_bgr = cv2.cvtColor(img_gray, cv2.COLOR_GRAY2BGR)

        img_alpha = img[..., 1]
        not_a = cv2.bitwise_not(img_alpha)
        not_a = cv2.cvtColor(not_a, cv2.COLOR_GRAY2BGR)

        new_img = cv2.bitwise_and(img_bgr, img_bgr, mask=img_alpha)
        new_img = cv2.add(new_img, not_a)
        return new_img

    @staticmethod
    def cvt_four_to_three(img: np.ndarray) -> np.ndarray:
        """RGBA → BGR"""
        r, g, b, a = cv2.split(img)
        new_img = cv2.merge((b, g, r))

        not_a = cv2.bitwise_not(a)
        not_a = cv2.cvtColor(not_a, cv2.COLOR_GRAY2BGR)

        new_img = cv2.bitwise_and(new_img, new_img, mask=a)

        mean_color = np.mean(new_img)
        if mean_color <= 0.0:
            new_img = cv2.add(new_img, not_a)
        else:
            new_img = cv2.bitwise_not(new_img)
        return new_img

    @staticmethod
    def verify_exist(file_path: Union[str, Path]):
        if not Path(file_path).exists():
            raise LoadImageError(f"{file_path} does not exist.")


class LoadImageError(Exception):
    pass


img_loader = LoadImage()


# from https://github.com/RapidAI/RapidTableDetection/blob/master/rapid_table_det/utils/visuallize.py
def visuallize(img, box, lt, rt, rb, lb):
    xmin, ymin, xmax, ymax = box
    draw_box = np.array([lt, rt, rb, lb]).reshape([-1, 2])
    cv2.circle(img, (int(lt[0]), int(lt[1])), 50, (255, 0, 0), 10)
    cv2.rectangle(img, (int(xmin), int(ymin)), (int(xmax), int(ymax)), (255, 0, 0), 10)
    cv2.polylines(
        img,
        [np.array(draw_box).astype(np.int32).reshape((-1, 1, 2))],
        True,
        color=(255, 0, 255),
        thickness=6,
    )
    return img


def extract_table_img(img, lt, rt, rb, lb):
    """
    根据四个角点进行透视变换，并提取出角点区域的图片。

    参数:
    img (numpy.ndarray): 输入图像
    lt (numpy.ndarray): 左上角坐标
    rt (numpy.ndarray): 右上角坐标
    lb (numpy.ndarray): 左下角坐标
    rb (numpy.ndarray): 右下角坐标

    返回:
    numpy.ndarray: 提取出的角点区域图片
    """
    # 源点坐标
    src_points = np.float32([lt, rt, lb, rb])

    # 目标点坐标
    width_a = np.sqrt(((rb[0] - lb[0]) ** 2) + ((rb[1] - lb[1]) ** 2))
    width_b = np.sqrt(((rt[0] - lt[0]) ** 2) + ((rt[1] - lt[1]) ** 2))
    max_width = max(int(width_a), int(width_b))

    height_a = np.sqrt(((rt[0] - rb[0]) ** 2) + ((rt[1] - rb[1]) ** 2))
    height_b = np.sqrt(((lt[0] - lb[0]) ** 2) + ((lt[1] - lb[1]) ** 2))
    max_height = max(int(height_a), int(height_b))

    dst_points = np.float32(
        [
            [0, 0],
            [max_width - 1, 0],
            [0, max_height - 1],
            [max_width - 1, max_height - 1],
        ]
    )

    # 获取透视变换矩阵
    M = cv2.getPerspectiveTransform(src_points, dst_points)

    # 应用透视变换
    warped = cv2.warpPerspective(img, M, (max_width, max_height))

    return warped

# from https://github.com/RapidAI/TableStructureRec/blob/main/wired_table_rec/utils/utils.py#L575
def insert_border_style(table_html_str: str):
    style_res = """<meta charset="UTF-8"><style>
    table {
        border-collapse: collapse;
        width: 100%;
    }
    th, td {
        border: 1px solid black;
        padding: 8px;
        text-align: center;
    }
    th {
        background-color: #f2f2f2;
    }
                </style>"""

    prefix_table, suffix_table = table_html_str.split("<body>")
    html_with_border = f"{prefix_table}{style_res}<body>{suffix_table}"
    return html_with_border


# 从图片中定位表格的位置
def callTableDectector(img_path):
    # img_path = f"chip.jpg"

    table_det = TableDetector(
        obj_model_path="ocr/RapidTableDet/rapid_table_paddle/models/obj_det_paddle",
        edge_model_path="ocr/RapidTableDet/rapid_table_paddle/models/edge_det_paddle",
        cls_model_path="ocr/RapidTableDet/rapid_table_paddle/models/cls_det_paddle",
        use_obj_det=True,
        use_edge_det=True,
        use_cls_det=True,
    )
    result, elapse = table_det(img_path)
    obj_det_elapse, edge_elapse, rotate_det_elapse = elapse
    print(
        f"obj_det_elapse:{obj_det_elapse}, edge_elapse={edge_elapse}, rotate_det_elapse={rotate_det_elapse}"
    )
    # 一张图片中可能有多个表格
    img = img_loader(img_path)
    file_name_with_ext = os.path.basename(img_path)
    file_name, file_ext = os.path.splitext(file_name_with_ext)
    out_dir = "rapid_table_det_paddle/outputs"
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    extract_img = img.copy()
    wrapped_imgs = []
    for i, res in enumerate(result):
        box = res["box"]
        lt, rt, rb, lb = res["lt"], res["rt"], res["rb"], res["lb"]
        # 带识别框和左上角方向位置
        img = visuallize(img, box, lt, rt, rb, lb)
        # 透视变换提取表格图片
        wrapped_img = extract_table_img(extract_img.copy(), lt, rt, rb, lb)
        wrapped_imgs.append(wrapped_img)
        cv2.imwrite(f"{out_dir}/{file_name}-extract-{i}.jpg", wrapped_img)
    cv2.imwrite(f"{out_dir}/{file_name}-visualize.jpg", img)
    return img, wrapped_imgs


# 调用 RapidOCR 识别表格中具体内容
def callRapidOCR(input_img_path):
    wired_input = WiredTableInput()
    lineless_input = LinelessTableInput()
    wired_engine = WiredTableRecognition(wired_input)
    lineless_engine = LinelessTableRecognition(lineless_input)
    # viser = VisTable()
    # 默认小yolo模型(0.1s)，可切换为精度更高yolox(0.25s),更快的qanything(0.07s)模型或paddle模型(0.03s)
    table_cls = TableCls()

    out_dir = "rapid_table_det_paddle/outputs"
    file_name_with_ext = os.path.basename(input_img_path)
    file_name, file_ext = os.path.splitext(file_name_with_ext)

    _, wrapped_table_img = callTableDectector(input_img_path)
    i = 0
    table_resultss = []
    for img in wrapped_table_img:
        cls, elasp = table_cls(img)
        if cls == "wired":
            table_engine = wired_engine
        else:
            table_engine = lineless_engine

        # 使用RapidOCR输入
        ocr_engine = RapidOCR()
        rapid_ocr_output = ocr_engine(img, return_word_box=True)
        ocr_result = list(
            zip(rapid_ocr_output.boxes, rapid_ocr_output.txts, rapid_ocr_output.scores)
        )
        table_results = table_engine(
            img, ocr_result=ocr_result
        )
        table_resultss.append(table_results)

        # # Save
        # img_path = f"{out_dir}/{file_name}-extract-{i}.jpg"
        # i += 1
        # save_dir = Path("outputs")
        # save_dir.mkdir(parents=True, exist_ok=True)
        
        # save_html_path = f"outputs/{Path(img_path).stem}.html"
        # save_drawed_path = f"outputs/{Path(img_path).stem}_table_vis{Path(img_path).suffix}"
        # save_logic_path = (
        #     f"outputs/{Path(img_path).stem}_table_vis_logic{Path(img_path).suffix}"
        # )

        # # Visualize table rec result
        # vis_imged = viser(
        #     img_path, table_results, save_html_path, save_drawed_path, save_logic_path
        # )
    return table_resultss


def html2excel(html_str, output_path):
    # 使用 BeautifulSoup 解析原始 HTML
    soup = BeautifulSoup(html_str, 'html.parser')
    table = soup.find('table')
    rows = table.find_all('tr')
    
    # 构建数据矩阵，用于处理 colspan 和 rowspan
    data_matrix = []
    max_cols = 0
    
    # 首先计算最大列数
    for tr in rows:
        cells = tr.find_all(['td', 'th'])
        cols = 0
        for cell in cells:
            colspan = int(cell.get('colspan', 1))
            cols += colspan
        max_cols = max(max_cols, cols)
    
    # 初始化数据矩阵
    for _ in rows:
        data_matrix.append([None] * max_cols)
    
    # 填充数据矩阵
    row_idx = 0
    for tr in rows:
        cells = tr.find_all(['td', 'th'])
        col_idx = 0
        
        # 找到下一个空位置
        while col_idx < max_cols and data_matrix[row_idx][col_idx] is not None:
            col_idx += 1
            
        for cell in cells:
            # 获取单元格属性
            colspan = int(cell.get('colspan', 1))
            rowspan = int(cell.get('rowspan', 1))
            value = cell.get_text(strip=True)
            
            # 找到下一个空位置
            while col_idx < max_cols and data_matrix[row_idx][col_idx] is not None:
                col_idx += 1
                
            # 填充数据矩阵，考虑 colspan 和 rowspan
            for r in range(row_idx, row_idx + rowspan):
                for c in range(col_idx, col_idx + colspan):
                    if r < len(data_matrix) and c < len(data_matrix[0]):
                        data_matrix[r][c] = value
            
            # 更新列索引
            col_idx += colspan
            
        row_idx += 1
    
    # 将数据转换为 DataFrame
    df = pandas.DataFrame(data_matrix)
    
    # 写入 Excel 并处理合并单元格
    with pandas.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, 'Sheet1', index=False, header=False)
        worksheet = writer.sheets["Sheet1"]
        
        # 处理合并单元格
        row_idx = 0
        for tr in rows:
            cells = tr.find_all(['td', 'th'])
            col_idx = 0
            
            # 找到下一个未处理的位置
            while col_idx < max_cols and worksheet.cell(row=row_idx+1, column=col_idx+1).value is None:
                col_idx += 1
                
            cell_idx = 0
            for cell in cells:
                colspan = int(cell.get('colspan', 1))
                rowspan = int(cell.get('rowspan', 1))
                
                # 找到下一个未处理的位置
                while col_idx < max_cols and worksheet.cell(row=row_idx+1, column=col_idx+1).value is None:
                    col_idx += 1
                
                start_row = row_idx + 1  # Excel 行号从 1 开始
                start_col = col_idx + 1  # Excel 列号从 1 开始
                end_row = start_row + rowspan - 1
                end_col = start_col + colspan - 1
                
                # 如果有合并，则进行合并操作
                if colspan > 1 or rowspan > 1:
                    # 清除合并区域中除左上角以外的值
                    for r in range(start_row, end_row + 1):
                        for c in range(start_col, end_col + 1):
                            if r != start_row or c != start_col:
                                worksheet.cell(row=r, column=c).value = None
                    
                    # 执行合并
                    worksheet.merge_cells(
                        start_row=start_row,
                        start_column=start_col,
                        end_row=end_row,
                        end_column=end_col
                    )
                    
                    # 居中对齐
                    cell_obj = worksheet.cell(row=start_row, column=start_col)
                    cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                
                # 更新列索引
                col_idx += colspan
                cell_idx += 1
                
            row_idx += 1

async def image2excel(input_img_path, output_path):
    table_resultss = callRapidOCR(input_img_path)

    i = 0
    ret_dict = {}
    for table_results in table_resultss:
        # Save
        save_dir = Path(output_path)
        save_html_path = f"{save_dir.parent}/{save_dir.stem}_{i}.html"
        save_excel_path = f"{save_dir.parent}/{save_dir.stem}_{i}.xlsx"

        html_with_border = insert_border_style(table_results.pred_html)
        with open(save_html_path, "w", encoding="utf-8") as f:
            f.write(html_with_border)

        # with open("outputs/table4-visualize.html", "r") as f:
        html2excel(html_with_border, save_excel_path)
        ret_dict[i] = {
            "html_content": html_with_border,
            "html_path": str(save_html_path),
            "excel_path": str(save_excel_path)
        }
        i += 1
    return ret_dict

def test_image2excel(input_img_path, output_path):
    table_resultss = callRapidOCR(input_img_path)

    ret_dict = {}
    i = 0
    for table_results in table_resultss:
        # Save
        save_dir = Path(output_path)
        save_html_path = f"{save_dir.parent}/{save_dir.stem}_{i}.html"
        save_excel_path = f"{save_dir.parent}/{save_dir.stem}_{i}.xlsx"

        html_with_border = insert_border_style(table_results.pred_html)
        with open(save_html_path, "w", encoding="utf-8") as f:
            f.write(html_with_border)

        # with open("outputs/table4-visualize.html", "r") as f:
        html2excel(html_with_border, save_excel_path)
        ret_dict[i] = {
            "html_content": html_with_border,
            "html_path": str(save_html_path),
            "excel_path": str(save_excel_path)
        }
        i += 1
    return ret_dict

if __name__ == "__main__":
    test_image2excel("doc.png", "doc.xlsx")